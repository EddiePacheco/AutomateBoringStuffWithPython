# This script takes two integers and a filename string as command line arguments.
# Where the first integer is N, the second integer is M, and the filename is the name of an Excel file.
# The script should insert M blank rows starting at row N in the Excel file.

# Step 1: Import the necessary modules.
# Step 2: Get the command line arguments.
# Step 3: Load the Excel file.
# Step 4: Insert M blank rows starting at row N.
# Step 5: Save the Excel file.

import openpyxl, sys, logging
logging.basicConfig(filename='blank_row_inserter_log.txt', level=logging.DEBUG, format='%(asctime)s - %(levelname)s - %(message)s')
from openpyxl.utils import get_column_letter # To convert column integer input to letter

# logging.disable(logging.CRITICAL) # Uncomment to disable logging

# Step 2: Get the command line arguments.
logging.debug('Checking command line arguments...')
if len(sys.argv) != 4:
    logging.error('Usage: python blank_row_inserter.py N M filename')
    sys.exit()

try:
    N = int(sys.argv[1])
    M = int(sys.argv[2])
    filename = str(sys.argv[3])
except ValueError:
    logging.error('N and M must be integers.')
    sys.exit()

# Step 3: Load the Excel file.
logging.debug('Opening workbook...')
try:
    wb = openpyxl.load_workbook(filename)
    sheet = wb.active
    logging.info(f'Workbook {filename} opened successfully.')
except FileNotFoundError:
    logging.error(f'File {filename} not found.')
    sys.exit()

# Step 4: Insert M blank rows starting at row N.
logging.debug('Inserting blank rows...')
for i in range(sheet.max_row, N - 1, -1): # Start from the last row and move up
    for j in range(1, sheet.max_column + 1):
        sheet.cell(row=i + M, column=j).value = sheet.cell(row=i, column=j).value # Move the value down
        sheet.cell(row=i, column=j).value = None # Clear the original cell

# Step 5: Save the Excel file.
logging.debug('Saving workbook...')
wb.save(f'{filename}')
logging.info(f'Workbook saved as {filename}.')
print(f'Blank rows inserted successfully in {filename}.')
logging.debug('Done.')