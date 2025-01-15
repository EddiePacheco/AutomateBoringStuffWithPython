# This script inverts the rows and columns of an Excel spreadsheet.
# Step 1: Import the necessary modules.
# Step 2: Get the command line arguments.
# Step 3: Load the Excel file.
# Step 4: Create a new Excel file.
# Step 5: Invert the rows and columns.
# Step 6: Save the new Excel file.

# Step 1: Import the necessary modules.
import openpyxl, sys, logging
logging.basicConfig(filename='spreadsheet_cell_inverter_log.txt', level=logging.DEBUG, format='%(asctime)s - %(levelname)s - %(message)s')
from openpyxl.utils import get_column_letter

# logging.disable(logging.CRITICAL) # Uncomment to disable logging

def log_and_print_debug(message):
    logging.debug(message)
    print(message)

def log_and_print_error(message):
    logging.error(message)
    print(message)

# Step 2: Get the command line arguments.
logging.debug('Checking command line arguments...')
if len(sys.argv) != 2:
    log_and_print_error('Usage: python spreadsheet_cell_inverter.py filename')
    sys.exit()

try:
    filename = str(sys.argv[1])
except ValueError:
    log_and_print_error('Filename must be a string.')
    sys.exit()

# Step 3: Load the Excel file.
log_and_print_debug('Opening workbook...')
try:
    wb = openpyxl.load_workbook(filename)
    sheet = wb.active
    log_and_print_debug(f'Workbook {filename} opened successfully.')
except FileNotFoundError:
    log_and_print_error(f'File {filename} not found.')
    sys.exit()

# Step 4: Create a new Excel file.
log_and_print_debug('Creating new workbook...')
new_wb = openpyxl.Workbook()
sheet_new = new_wb.active

# Step 5: Invert the rows and columns.
log_and_print_debug('Inverting rows and columns...')
for i in range(1, sheet.max_row + 1):
    for j in range(1, sheet.max_column + 1):
        sheet_new.cell(row=j, column=i).value = sheet.cell(row=i, column=j).value

# Step 6: Save the new Excel file.
log_and_print_debug('Saving workbook...')
new_wb.save(f'inverted_{filename}')
log_and_print_debug(f'Workbook saved as inverted_{filename}.\n')
log_and_print_debug(f'Rows and columns inverted successfully in {filename}.\n')
log_and_print_debug('Done.')
