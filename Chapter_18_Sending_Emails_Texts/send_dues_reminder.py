#! python3
# send_dues_reminder.py - Sends emails based on payment status in spreadsheet.

# Read data from an Excel spreadsheet.
# Find all members who have not paid dues for the latest month.
# Find their email addresses and send them a personalized reminder email.

import openpyxl, smtplib, sys

# Open the spreadsheet and get the latest dues status.
wb = openpyxl.load_workbook('duesRecords.xlsx')

sheet = wb.get_sheet_by_name('Sheet1')

last_col = sheet.max_column
latest_month = sheet.cell(row=1, column=last_col).value

# Check each member's payment status.
unpaid_members = {}
for i in range(2, sheet.max_row + 1):
    payment = sheet.cell(row=i, column=last_col).value
    if payment != 'paid':
        name = sheet.cell(row=i, column=1).value
        email = sheet.cell(row=i, column=2).value
        unpaid_members[name] = email # Add the key-value pair to the dictionary.

# Log in to email account.
smtp_obj = smtplib.SMTP('smtp.gmail.com', 587)

smtp_obj.ehlo()

smtp_obj.starttls()

user_email = input('Enter your email: ')
user_password = input('Enter your password: ')
smtp_obj.login(user_email, user_password)

# Send out reminder emails.
for name, email in unpaid_members.items():
    body = (
        "Subject: %s dues unpaid.\nDear %s,\nRecords show that you have not paid "
        "dues for %s. Please make this payment as soon as possible. Thank you!" 
        % (latest_month, name, latest_month)
    )
    print('Sending email to %s...' % email)

    sendmail_status = smtp_obj.sendmail(user_email, email, body)

    if not sendmail_status:
        print('There was a problem sending email to %s: %s' % (email, sendmail_status))

smtp_obj.quit()

